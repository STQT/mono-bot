"""
Admin configuration for core models.
"""
import zipfile
import os
from django.contrib import admin
from rangefilter.filters import DateTimeRangeFilterBuilder, DateRangeFilterBuilder


class NoDeleteAdminMixin:
    """Отключает удаление везде в админке (включая action «Удалить выбранные»)."""
    def has_delete_permission(self, request, obj=None):
        return False

    def get_actions(self, request):
        actions = super().get_actions(request)
        if 'delete_selected' in actions:
            del actions['delete_selected']
        return actions
from django.http import HttpResponse
from django.utils.html import format_html
from django.urls import path
from django.shortcuts import render, redirect
from django.template.response import TemplateResponse
from django.contrib import messages
from django.conf import settings
from django.db import models
from simple_history.admin import SimpleHistoryAdmin
from .models import (
    TelegramUser, QRCode, QRCodeScanAttempt,
    Gift, GiftRedemption, BroadcastMessage, RegionMessageLog, Promotion, QRCodeGeneration, PrivacyPolicy, AdminContactSettings, VideoInstruction, SmartUPId
)
from .utils import generate_qr_code_image, generate_qr_codes_batch


class SmartUPFilter(admin.SimpleListFilter):
    """Фильтр: пользователи со SmartUP ID и без."""
    title = 'SmartUP'
    parameter_name = 'has_smartup'

    def lookups(self, request, model_admin):
        return (
            ('yes', 'Со SmartUP'),
            ('no', 'Без SmartUP'),
        )

    def queryset(self, request, queryset):
        if self.value() == 'yes':
            return queryset.filter(smartup_id__isnull=False).exclude(smartup_id=0)
        if self.value() == 'no':
            return queryset.filter(models.Q(smartup_id__isnull=True) | models.Q(smartup_id=0))
        return queryset


class ScannedQRCodeInline(admin.TabularInline):
    """Инлайн: успешно отсканированные промокоды пользователем (readonly)."""
    model = QRCode
    fk_name = 'scanned_by'
    extra = 0
    can_delete = False
    can_add = False
    max_num = 0
    readonly_fields = ['serial_number', 'code_type', 'points', 'scanned_at']
    fields = ['serial_number', 'code_type', 'points', 'scanned_at']
    ordering = ['-scanned_at']
    verbose_name = 'Отсканированный промокод'
    verbose_name_plural = 'Отсканированные промокоды'

    def get_queryset(self, request):
        return super().get_queryset(request).filter(is_scanned=True)


@admin.register(TelegramUser)
class TelegramUserAdmin(NoDeleteAdminMixin, SimpleHistoryAdmin):
    """Админка для пользователей Telegram."""
    inlines = [ScannedQRCodeInline]
    list_display = [
        'user_display', 'phone_number', 'region_display', 'district_display', 
        'user_type_badge', 'points_display', 'language_badge', 'status_badge', 'created_at', 'send_message_button'
    ]
    list_filter = [
        'user_type', 'is_active', 'language', 'region', 'district',
        SmartUPFilter,
        ('created_at', DateTimeRangeFilterBuilder(title='Дата регистрации (диапазон)')),
    ]
    search_fields = ['telegram_id', 'username', 'first_name', 'phone_number']
    readonly_fields = [
        'telegram_id', 'created_at', 'updated_at',
        'last_message_sent_at', 'blocked_bot_at', 'region', 'district',
        'points', 'points_display',
    ]
    ordering = ['region', 'district', '-created_at']
    actions = ['send_personal_message_action', 'update_locations_action', 'change_user_type_to_electrician', 'change_user_type_to_seller']
    list_per_page = 50
    date_hierarchy = 'created_at'
    change_list_template = 'admin/core/telegramuser/change_list.html'

    class Media:
        css = {'all': ('core_admin/css/changelist_filters.css',)}
        js = ('core_admin/js/changelist_filters.js',)

    def changelist_view(self, request, extra_context=None):
        from django.urls import reverse
        extra_context = extra_context or {}
        if request.user.has_perm('core.send_region_messages'):
            extra_context['send_region_message_url'] = reverse('admin:core_telegramuser_send_region_message')
        return super().changelist_view(request, extra_context)

    def user_display(self, obj):
        """Отображает пользователя с иконкой и ссылкой."""
        icon = "⚡" if obj.user_type == 'electrician' else "🛒"
        name = obj.first_name or "Пользователь"
        username = f"@{obj.username}" if obj.username else ""
        return format_html(
            '<span style="font-size: 18px;">{}</span> <strong>{}</strong> <span style="color: #718096;">{}</span><br>'
            '<span style="color: #718096; font-size: 12px;">ID: {}</span>',
            icon, name, username, obj.telegram_id
        )
    user_display.short_description = 'Пользователь'
    user_display.admin_order_field = 'first_name'
    
    def user_type_badge(self, obj):
        """Отображает тип пользователя с цветным badge."""
        if obj.user_type == 'electrician':
            return format_html(
                '<span style="background: #fef3c7; color: #92400e; padding: 4px 12px; border-radius: 12px; '
                'font-size: 12px; font-weight: 600;">⚡ Elektrik</span>'
            )
        elif obj.user_type == 'seller':
            return format_html(
                '<span style="background: #dbeafe; color: #1e40af; padding: 4px 12px; border-radius: 12px; '
                'font-size: 12px; font-weight: 600;">🛒 Sotuvchi</span>'
            )
        return '-'
    user_type_badge.short_description = 'Тип'
    user_type_badge.admin_order_field = 'user_type'
    
    def points_display(self, obj):
        """Отображает баллы с цветом (вычисляются динамически: промокоды − активные заказы)."""
        if obj is None:
            return '-'
        try:
            calculated = obj.calculate_points()
        except Exception:
            calculated = obj.points
        points_formatted = f"{calculated:,}".replace(",", " ")
        return format_html(
            '<span style="color: #667eea; font-weight: 700; font-size: 16px;">{} баллов</span>',
            points_formatted
        )
    points_display.short_description = 'Баллы (промокоды − заказы)'
    points_display.admin_order_field = 'points'
    
    def language_badge(self, obj):
        """Отображает язык с цветным badge."""
        colors = {
            'uz_latin': ('#dbeafe', '#1e40af', '🇺🇿'),
            'ru': ('#fee2e2', '#991b1b', '🇷🇺'),
        }
        bg, text, flag = colors.get(obj.language, ('#f3f4f6', '#374151', '🌐'))
        label = dict(obj._meta.get_field('language').choices).get(obj.language, obj.language)
        return format_html(
            '<span style="background: {}; color: {}; padding: 4px 12px; border-radius: 12px; '
            'font-size: 12px; font-weight: 600;">{} {}</span>',
            bg, text, flag, label.split('(')[0].strip()
        )
    language_badge.short_description = 'Язык'
    language_badge.admin_order_field = 'language'
    
    def status_badge(self, obj):
        """Отображает статус активности."""
        if obj.is_active:
            return format_html(
                '<span style="background: #d4edda; color: #155724; padding: 4px 12px; border-radius: 12px; '
                'font-size: 12px; font-weight: 600;">✅ Активен</span>'
            )
        else:
            return format_html(
                '<span style="background: #f8d7da; color: #721c24; padding: 4px 12px; border-radius: 12px; '
                'font-size: 12px; font-weight: 600;">❌ Неактивен</span>'
            )
    status_badge.short_description = 'Статус'
    status_badge.admin_order_field = 'is_active'
    
    def region_display(self, obj):
        """Отображает область пользователя."""
        region_name = obj.get_region_display('ru')
        if region_name:
            return format_html(
                '<span style="background: #e0e7ff; color: #3730a3; padding: 4px 12px; border-radius: 12px; '
                'font-size: 12px; font-weight: 600;">📍 {}</span>',
                region_name
            )
        elif obj.latitude and obj.longitude:
            return format_html(
                '<span style="color: #718096; font-size: 12px;">Не определено</span>'
            )
        else:
            return format_html(
                '<span style="color: #cbd5e0; font-size: 12px;">-</span>'
            )
    region_display.short_description = 'Область'
    region_display.admin_order_field = 'region'
    
    def district_display(self, obj):
        """Отображает район пользователя."""
        district_name = obj.get_district_display('ru')
        if district_name:
            return format_html(
                '<span style="background: #fef3c7; color: #92400e; padding: 4px 12px; border-radius: 12px; '
                'font-size: 12px; font-weight: 600;">🏘️ {}</span>',
                district_name
            )
        elif obj.latitude and obj.longitude:
            return format_html(
                '<span style="color: #718096; font-size: 12px;">Не определено</span>'
            )
        else:
            return format_html(
                '<span style="color: #cbd5e0; font-size: 12px;">-</span>'
            )
    district_display.short_description = 'Район'
    district_display.admin_order_field = 'district'
    
    fieldsets = (
        ('Основная информация', {
            'fields': ('telegram_id', 'username', 'first_name', 'last_name')
        }),
        ('Контактные данные', {
            'fields': ('phone_number', 'latitude', 'longitude', 'region', 'district')
        }),
        ('Тип и баллы', {
            'fields': ('user_type', 'points_display', 'points', 'smartup_id'),
            'description': 'Баллы рассчитываются по отсканированным промокодам минус активные заказы на подарки.',
        }),
        ('Настройки', {
            'fields': ('language',)
        }),
        ('Активность', {
            'fields': ('is_active', 'last_message_sent_at', 'blocked_bot_at')
        }),
        ('Даты', {
            'fields': ('created_at', 'updated_at')
        }),
    )
    
    def get_readonly_fields(self, request, obj=None):
        """Управляет readonly полями в зависимости от роли пользователя."""
        readonly = list(super().get_readonly_fields(request, obj))
        
        # Если пользователь имеет пермишн call center и не является superuser
        if not request.user.is_superuser and request.user.has_perm('core.change_user_type_call_center'):
            # Получаем только конкретные поля модели (не обратные связи)
            model_fields = [
                f.name for f in TelegramUser._meta.get_fields() 
                if isinstance(f, models.Field) and hasattr(f, 'name')
            ]
            # Исключаем user_type - это единственное поле, которое Call Center может менять
            fields_to_make_readonly = [f for f in model_fields if f != 'user_type']
            
            # Добавляем все поля в readonly, кроме user_type
            for field in fields_to_make_readonly:
                if field not in readonly:
                    readonly.append(field)
        # Для обычных админов (не superuser и не Call Center) user_type доступен для редактирования
        # Он не в списке readonly_fields, поэтому будет доступен по умолчанию
        
        return readonly
    
    def send_personal_message_action(self, request, queryset):
        """Действие для отправки персонального сообщения."""
        from django.shortcuts import render
        from django import forms
        
        class MessageForm(forms.Form):
            message = forms.CharField(widget=forms.Textarea, label='Текст сообщения')
            parse_mode = forms.ChoiceField(
                choices=[('', 'Без форматирования'), ('HTML', 'HTML'), ('Markdown', 'Markdown')],
                required=False,
                label='Режим парсинга'
            )
        
        if request.method == 'POST':
            form = MessageForm(request.POST)
            if form.is_valid():
                message_text = form.cleaned_data['message']
                parse_mode = form.cleaned_data['parse_mode'] or None
                
                import asyncio
                from django.conf import settings
                from aiogram import Bot
                from core.messaging import send_personal_message
                
                async def send_messages():
                    bot = Bot(token=settings.TELEGRAM_BOT_TOKEN)
                    try:
                        sent = 0
                        failed = 0
                        for user in queryset:
                            success, error = await send_personal_message(
                                bot=bot,
                                telegram_id=user.telegram_id,
                                text=message_text,
                                parse_mode=parse_mode
                            )
                            if success:
                                sent += 1
                            else:
                                failed += 1
                        return sent, failed
                    finally:
                        await bot.session.close()
                
                sent, failed = asyncio.run(send_messages())
                self.message_user(
                    request,
                    f'Отправлено: {sent}, Ошибок: {failed}',
                    level=messages.SUCCESS if failed == 0 else messages.WARNING
                )
                return redirect('admin:core_telegramuser_changelist')
        else:
            form = MessageForm()
        
        from django.template.response import TemplateResponse
        
        context = {
            **self.admin_site.each_context(request),
            'form': form,
            'users': queryset,
            'title': 'Отправить сообщение пользователям',
            'opts': self.model._meta,
            'has_view_permission': True,
            'has_add_permission': False,
            'has_change_permission': False,
            'has_delete_permission': False,
        }
        
        return TemplateResponse(request, 'admin/core/telegramuser/send_message.html', context)
    send_personal_message_action.short_description = 'Отправить персональное сообщение выбранным пользователям'
    
    def get_search_results(self, request, queryset, search_term):
        """Кастомный поиск с поддержкой поиска по последним 4 цифрам номера телефона."""
        queryset, use_distinct = super().get_search_results(request, queryset, search_term)
        
        # Если поисковый запрос состоит из 4 цифр, ищем по последним 4 цифрам номера телефона
        if search_term and len(search_term) == 4 and search_term.isdigit():
            from django.db.models import Q, CharField
            from django.db.models.functions import Right, Replace
            
            # Ищем номера телефонов, которые заканчиваются на эти 4 цифры
            # Учитываем разные форматы номеров (с пробелами, дефисами, плюсами и т.д.)
            # Используем регулярное выражение для поиска номеров, заканчивающихся на эти 4 цифры
            # Паттерн ищет номера, которые заканчиваются на эти 4 цифры (возможно с разделителями)
            phone_pattern = rf'{search_term}$'
            
            # Прямой поиск по окончанию
            phone_query = Q(phone_number__endswith=search_term)
            
            # Поиск с учетом разделителей перед последними 4 цифрами
            # Ищем паттерны типа: -4567,  4567, (4567) и т.д.
            # Регулярное выражение ищет номера, которые заканчиваются на эти 4 цифры
            # с возможными разделителями (пробелы, дефисы, скобки и т.д.) перед ними
            phone_query |= Q(phone_number__iregex=rf'[\s\-\(\)\.]*{search_term}$')
            
            phone_results = self.model.objects.filter(phone_query)
            
            # Объединяем результаты
            queryset = queryset | phone_results
            use_distinct = True
        
        return queryset, use_distinct
    
    
    def update_locations_action(self, request, queryset):
        """Действие для обновления локаций выбранных пользователей."""
        updated = 0
        for user in queryset:
            if user.latitude is not None and user.longitude is not None:
                user.update_location()
                user.save(update_fields=['region', 'district'])
                updated += 1
        
        self.message_user(
            request,
            f'Обновлено локаций: {updated} из {queryset.count()}',
            messages.SUCCESS
        )
    update_locations_action.short_description = 'Обновить локации (область и район)'
    
    def change_user_type_to_electrician(self, request, queryset):
        """Массовое изменение типа пользователя на Электрик."""
        updated = queryset.update(user_type='electrician')
        self.message_user(
            request,
            f'Тип пользователя изменен на "Электрик" для {updated} пользователей.',
            messages.SUCCESS
        )
    change_user_type_to_electrician.short_description = 'Изменить тип на: ⚡ Электрик'
    
    def change_user_type_to_seller(self, request, queryset):
        """Массовое изменение типа пользователя на Продавец (Предприниматель)."""
        updated = queryset.update(user_type='seller')
        self.message_user(
            request,
            f'Тип пользователя изменен на "Продавец (Предприниматель)" для {updated} пользователей.',
            messages.SUCCESS
        )
    change_user_type_to_seller.short_description = 'Изменить тип на: 🛒 Продавец (Предприниматель)'
    
    def send_message_button(self, obj):
        """Кнопка отправки сообщения в списке."""
        from django.urls import reverse
        url = reverse('admin:core_telegramuser_send_single_message', args=[obj.pk])
        return format_html(
            '<a href="{}" style="background: #667eea; color: white; padding: 6px 12px; '
            'border-radius: 4px; text-decoration: none; white-space: nowrap; font-size: 12px;">💬 Написать</a>',
            url
        )
    send_message_button.short_description = 'Сообщение'
    
    def get_urls(self):
        """Добавляет кастомные URL."""
        urls = super().get_urls()
        custom_urls = [
            path('<int:user_id>/send_message/', self.admin_site.admin_view(self.send_single_message_view), name='core_telegramuser_send_single_message'),
            path('send_region_message/', self.admin_site.admin_view(self.send_region_message_view), name='core_telegramuser_send_region_message'),
        ]
        return custom_urls + urls
    
    def send_single_message_view(self, request, user_id):
        """Страница отправки сообщения конкретному пользователю."""
        from django import forms
        
        user = TelegramUser.objects.get(pk=user_id)
        
        class MessageForm(forms.Form):
            message = forms.CharField(widget=forms.Textarea, label='Текст сообщения')
            parse_mode = forms.ChoiceField(
                choices=[('', 'Без форматирования'), ('HTML', 'HTML'), ('Markdown', 'Markdown')],
                required=False,
                label='Режим парсинга'
            )
        
        if request.method == 'POST':
            form = MessageForm(request.POST)
            if form.is_valid():
                message_text = form.cleaned_data['message']
                parse_mode = form.cleaned_data['parse_mode'] or None
                
                import asyncio
                from core.messaging import send_personal_message
                
                async def send():
                    from aiogram import Bot
                    bot = Bot(token=settings.TELEGRAM_BOT_TOKEN)
                    try:
                        return await send_personal_message(bot=bot, telegram_id=user.telegram_id, text=message_text, parse_mode=parse_mode)
                    finally:
                        await bot.session.close()
                
                success, error = asyncio.run(send())
                if success:
                    self.message_user(request, f'Сообщение отправлено пользователю {user}', messages.SUCCESS)
                else:
                    self.message_user(request, f'Ошибка: {error}', messages.ERROR)
                return redirect('admin:core_telegramuser_changelist')
        else:
            form = MessageForm()
        
        context = {
            **self.admin_site.each_context(request),
            'form': form,
            'users': TelegramUser.objects.filter(pk=user_id),
            'title': f'Отправить сообщение: {user}',
            'opts': self.model._meta,
            'has_view_permission': True,
            'has_add_permission': False,
            'has_change_permission': False,
            'has_delete_permission': False,
        }
        return TemplateResponse(request, 'admin/core/telegramuser/send_message.html', context)

    def send_region_message_view(self, request):
        """Страница отправки сообщения по области (с фото, форматированием, ссылками)."""
        from django import forms
        from django.core.exceptions import PermissionDenied
        from core.regions import get_region_by_coordinates, get_all_regions

        if not request.user.has_perm('core.send_region_messages'):
            raise PermissionDenied

        region_choices = [
            ('', '--- Выберите область ---'),
            ('all', 'Все регионы'),
        ] + list(get_all_regions('ru'))

        class RegionMessageForm(forms.Form):
            region = forms.ChoiceField(choices=region_choices, required=True, label='Область')
            message = forms.CharField(widget=forms.Textarea(attrs={'rows': 8}), label='Текст сообщения', required=False)
            image = forms.ImageField(required=False, label='Фото (опционально)')
            user_type_filter = forms.ChoiceField(
                choices=[('', 'Все'), ('electrician', 'Электрики'), ('seller', 'Продавцы')],
                required=False,
                label='Тип пользователя'
            )
            language_filter = forms.ChoiceField(
                choices=[('', 'Все языки'), ('uz_latin', "O'zbek (Lotin)"), ('ru', 'Русский')],
                required=False,
                label='Язык пользователя'
            )

        if request.method == 'POST':
            form = RegionMessageForm(request.POST, request.FILES)
            if form.is_valid():
                region_code = form.cleaned_data['region']
                message_text = form.cleaned_data.get('message') or ''
                image_file = form.cleaned_data.get('image')
                user_type_filter = form.cleaned_data['user_type_filter'] or None
                language_filter = form.cleaned_data.get('language_filter') or None

                users_qs = TelegramUser.objects.filter(
                    latitude__isnull=False,
                    longitude__isnull=False,
                    is_active=True
                )
                if user_type_filter:
                    users_qs = users_qs.filter(user_type=user_type_filter)
                if language_filter:
                    users_qs = users_qs.filter(language=language_filter)

                users = list(users_qs)
                if region_code == 'all':
                    filtered = users
                else:
                    filtered = [
                        u for u in users
                        if get_region_by_coordinates(u.latitude, u.longitude) == region_code
                    ]

                if not filtered:
                    msg = 'Нет пользователей с координатами.' if region_code == 'all' else 'В выбранной области нет пользователей с координатами.'
                    self.message_user(request, msg, messages.WARNING)
                else:
                    from core.tasks import send_region_message_task, REGION_MESSAGE_ASYNC_THRESHOLD
                    from core.messaging import TELEGRAM_MESSAGE_DELAY

                    n = len(filtered)
                    # Большая рассылка — в фоне (нет таймаута админки, соблюдаются лимиты Telegram)
                    if n > REGION_MESSAGE_ASYNC_THRESHOLD:
                        import os
                        import uuid
                        from django.core.files.storage import default_storage
                        from django.core.files.base import ContentFile

                        image_storage_path = ''
                        if image_file:
                            ext = os.path.splitext(image_file.name)[1] or '.jpg'
                            name = f'region_messages/{uuid.uuid4().hex}{ext}'
                            default_storage.save(name, ContentFile(image_file.read()))
                            image_storage_path = name

                        log = RegionMessageLog.objects.create(
                            region_code=region_code,
                            user_type_filter=user_type_filter,
                            language_filter=language_filter,
                            total=n,
                            status='running',
                            initiated_by=request.user,
                        )
                        send_region_message_task.delay(
                            log_id=log.id,
                            region_code=region_code,
                            message_text=message_text,
                            image_storage_path=image_storage_path,
                            user_type_filter=user_type_filter,
                            language_filter=language_filter,
                        )
                        self.message_user(
                            request,
                            f'Рассылка по области запущена в фоне ({n} пользователей). '
                            'Результаты — в разделе «Логи рассылок по областям».',
                            messages.SUCCESS,
                        )
                        return redirect('admin:core_regionmessagelog_changelist')

                    # Небольшая рассылка — сразу в этом запросе
                    import asyncio
                    import tempfile
                    import os

                    photo_path = None
                    if image_file:
                        ext = os.path.splitext(image_file.name)[1] or '.jpg'
                        with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp:
                            for chunk in image_file.chunks():
                                tmp.write(chunk)
                            photo_path = tmp.name

                    async def send_all():
                        from aiogram import Bot
                        bot = Bot(token=settings.TELEGRAM_BOT_TOKEN)
                        sent, failed = 0, 0
                        try:
                            for i, user in enumerate(filtered):
                                from core.messaging import send_message_to_user
                                success, err = await send_message_to_user(
                                    bot=bot, user=user, text=message_text,
                                    parse_mode='HTML', photo_path=photo_path
                                )
                                if success:
                                    sent += 1
                                else:
                                    failed += 1
                                if i < len(filtered) - 1:
                                    await asyncio.sleep(TELEGRAM_MESSAGE_DELAY)
                            return sent, failed
                        finally:
                            await bot.session.close()
                            if photo_path and os.path.exists(photo_path):
                                try:
                                    os.unlink(photo_path)
                                except OSError:
                                    pass

                    sent, failed = asyncio.run(send_all())
                    self.message_user(
                        request,
                        f'Отправлено: {sent}, ошибок: {failed} (всего {len(filtered)} пользователей)',
                        messages.SUCCESS
                    )
                    return redirect('admin:core_telegramuser_changelist')
        else:
            form = RegionMessageForm()

        context = {
            **self.admin_site.each_context(request),
            'form': form,
            'title': 'Отправить сообщение по области',
            'opts': self.model._meta,
        }
        return TemplateResponse(request, 'admin/core/telegramuser/send_region_message.html', context)


class QRCodeScanAttemptInline(admin.TabularInline):
    """Инлайн для попыток сканирования."""
    model = QRCodeScanAttempt
    extra = 0
    readonly_fields = ['user', 'attempted_at', 'is_successful']
    can_delete = False


@admin.register(QRCode)
class QRCodeAdmin(NoDeleteAdminMixin, SimpleHistoryAdmin):
    """Админка для QR-кодов (только просмотр)."""
    list_display = [
        'qr_display', 'code_type_badge', 'points_display', 
        'status_badge', 'scanned_by_display', 'generated_at'
    ]
    list_filter = [
        'code_type', 'is_scanned',
    ]
    search_fields = ['code', 'hash_code', 'serial_number']
    readonly_fields = [
        'code', 'code_type', 'hash_code', 'serial_number',
        'points', 'generated_at', 'scanned_at', 'scanned_by', 'is_scanned'
    ]
    ordering = ['-generated_at']
    inlines = [QRCodeScanAttemptInline]
    list_per_page = 50
    date_hierarchy = 'generated_at'
    
    def has_view_permission(self, request, obj=None):
        """Проверяет права доступа к просмотру QR-кода."""
        # Superuser всегда имеет доступ
        if request.user.is_superuser:
            return True
        
        # Проверяем custom permission
        if request.user.has_perm('core.view_qrcode_detail'):
            return True
        
        return False
    
    def get_list_display_links(self, request, list_display):
        """Скрывает ссылки на детальный просмотр для пользователей без permission."""
        if not self.has_view_permission(request):
            # Если нет доступа к просмотру, не показываем ссылки
            return (None,)
        # По умолчанию Django использует первый элемент list_display как ссылку
        return super().get_list_display_links(request, list_display)
    
    def get_fields(self, request, obj=None):
        """Возвращает список полей для отображения, скрывая code и hash_code для неиспользованных QR-кодов."""
        fields = list(super().get_fields(request, obj))
        
        # Всегда скрываем image_path
        if 'image_path' in fields:
            fields.remove('image_path')
        
        # Скрываем code и hash_code для неиспользованных QR-кодов (безопасность)
        if obj and not obj.is_scanned:
            if 'code' in fields:
                fields.remove('code')
            if 'hash_code' in fields:
                fields.remove('hash_code')
            # Добавляем информационное поле вместо code
            if 'security_notice' not in fields:
                # Вставляем после code_type или в начало
                try:
                    code_type_index = fields.index('code_type')
                    fields.insert(code_type_index + 1, 'security_notice')
                except ValueError:
                    fields.insert(0, 'security_notice')
        
        # Если пользователь не имеет прав на просмотр деталей, заменяем code на masked_code_display
        elif obj and not self.has_view_permission(request, obj):
            # Сохраняем индекс code перед удалением
            code_index = None
            if 'code' in fields:
                code_index = fields.index('code')
                fields.remove('code')
            if 'hash_code' in fields:
                fields.remove('hash_code')
            if 'masked_code_display' not in fields:
                # Вставляем masked_code_display на место code
                if code_index is not None:
                    fields.insert(code_index, 'masked_code_display')
                else:
                    # Если code не найден, просто добавляем в начало
                    fields.insert(0, 'masked_code_display')
        
        return fields
    
    def get_readonly_fields(self, request, obj=None):
        """Возвращает список readonly полей, добавляя информационное поле для неиспользованных QR-кодов."""
        readonly = list(super().get_readonly_fields(request, obj))
        
        # Для неиспользованных QR-кодов добавляем информационное поле
        if obj and not obj.is_scanned:
            # Убираем code и hash_code из readonly, так как мы их скрываем
            if 'code' in readonly:
                readonly.remove('code')
            if 'hash_code' in readonly:
                readonly.remove('hash_code')
            # Добавляем security_notice
            if 'security_notice' not in readonly:
                readonly.append('security_notice')
        
        # Если пользователь не имеет прав на просмотр деталей, маскируем код
        elif obj and not self.has_view_permission(request, obj):
            # Убираем code из readonly, так как мы заменим его на masked_code
            if 'code' in readonly:
                readonly.remove('code')
            if 'hash_code' in readonly:
                readonly.remove('hash_code')
            # Добавляем masked_code вместо code
            if 'masked_code_display' not in readonly:
                readonly.append('masked_code_display')
        
        return readonly
    
    def masked_code_display(self, obj):
        """Отображает замаскированный код для пользователей без прав."""
        if obj:
            masked = self.masked_code(obj)
            return format_html(
                '<div style="font-family: monospace; font-size: 14px; color: #333;">'
                '<strong>{}</strong></div>',
                masked
            )
        return '-'
    masked_code_display.short_description = 'Code'
    
    def security_notice(self, obj):
        """Информационное сообщение о безопасности для неиспользованных QR-кодов."""
        if obj and not obj.is_scanned:
            return format_html(
                '<div style="background: #fff3cd; border: 1px solid #ffc107; border-radius: 8px; '
                'padding: 15px; margin: 10px 0;">'
                '<div style="display: flex; align-items: center; gap: 10px; margin-bottom: 10px;">'
                '<span style="font-size: 20px;">🔒</span>'
                '<strong style="color: #856404; font-size: 14px;">Информация о безопасности</strong>'
                '</div>'
                '<p style="margin: 0; color: #856404; font-size: 13px; line-height: 1.5;">'
                'Код QR-кода скрыт для безопасности, так как QR-код еще не использован. '
                'После сканирования QR-кода пользователем код будет доступен для просмотра.'
                '</p>'
                '<p style="margin: 10px 0 0 0; color: #856404; font-size: 12px;">'
                '<strong>Серийный номер:</strong> {}</p>'
                '</div>',
                obj.serial_number if obj else '-'
            )
        return '-'
    security_notice.short_description = 'Информация'
    
    def change_view(self, request, object_id, form_url='', extra_context=None):
        """Переопределяем детальный просмотр для проверки прав доступа и маскирования кода."""
        from django.template.response import TemplateResponse
        
        obj = self.get_object(request, object_id)
        
        # Проверяем права доступа
        if not self.has_view_permission(request, obj):
            # Если нет доступа, показываем кастомный шаблон с сообщением
            extra_context = extra_context or {}
            extra_context['no_access'] = True
            extra_context['is_superuser'] = request.user.is_superuser
            extra_context['has_permission'] = request.user.has_perm('core.view_qrcode_detail')
            extra_context['title'] = 'Доступ запрещен'
            extra_context['opts'] = self.model._meta
            extra_context['has_view_permission'] = False
            extra_context['has_add_permission'] = False
            extra_context['has_change_permission'] = False
            extra_context['has_delete_permission'] = False
            
            return TemplateResponse(
                request,
                'admin/core/qrcode/no_access.html',
                extra_context,
                status=403
            )
        
        return super().change_view(request, object_id, form_url, extra_context)
    
    def qr_display(self, obj):
        """Отображает QR-код с серийным номером."""
        return format_html(
            '<div style="line-height: 1.6;">'
            '<strong style="font-size: 16px;">📱 #{}</strong><br>'
            '<span style="color: #718096; font-size: 12px; font-family: monospace;">{}</span>',
            obj.serial_number,
            self.masked_code(obj)
        )
    qr_display.short_description = 'QR-код'
    qr_display.admin_order_field = 'serial_number'
    
    def code_type_badge(self, obj):
        """Отображает тип кода."""
        if obj.code_type == 'electrician':
            return format_html(
                '<span style="background: #fef3c7; color: #92400e; padding: 4px 12px; border-radius: 12px; '
                'font-size: 12px; font-weight: 600;">⚡ E-</span>'
            )
        elif obj.code_type == 'seller':
            return format_html(
                '<span style="background: #dbeafe; color: #1e40af; padding: 4px 12px; border-radius: 12px; '
                'font-size: 12px; font-weight: 600;">🛒 D-</span>'
            )
        return '-'
    code_type_badge.short_description = 'Тип'
    code_type_badge.admin_order_field = 'code_type'
    
    def points_display(self, obj):
        """Отображает баллы."""
        points_formatted = f"{obj.points:,}"
        return format_html(
            '<span style="color: #667eea; font-weight: 700; font-size: 16px;">{}</span>',
            points_formatted
        )
    points_display.short_description = 'Баллы'
    points_display.admin_order_field = 'points'
    
    def status_badge(self, obj):
        """Отображает статус сканирования."""
        if obj.is_scanned:
            return format_html(
                '<span style="background: #d4edda; color: #155724; padding: 4px 12px; border-radius: 12px; '
                'font-size: 12px; font-weight: 600;">✅ Использован</span>'
            )
        else:
            return format_html(
                '<span style="background: #fff3cd; color: #856404; padding: 4px 12px; border-radius: 12px; '
                'font-size: 12px; font-weight: 600;">⏳ Не использован</span>'
            )
    status_badge.short_description = 'Статус'
    status_badge.admin_order_field = 'is_scanned'
    
    def changelist_view(self, request, extra_context=None):
        """Добавляет кнопку для генерации QR-кодов и информацию о доступе."""
        extra_context = extra_context or {}
        extra_context['show_generate_button'] = True
        extra_context['has_view_permission'] = self.has_view_permission(request)
        return super().changelist_view(request, extra_context=extra_context)
    
    def has_add_permission(self, request):
        """Отключаем добавление через админку."""
        return False
    
    def has_change_permission(self, request, obj=None):
        """Отключаем редактирование."""
        return False
    
    def masked_code(self, obj):
        """Маскирует часть кода для отображения."""
        if len(obj.code) > 5:
            # Для коротких кодов: E-ABC123 -> E-AB***3
            prefix = obj.code[:3]  # E- или D- + первый символ
            suffix = obj.code[-1]   # Последний символ
            masked = '*' * max(1, len(obj.code) - 4)
            return f"{prefix}{masked}{suffix}"
        return obj.code
    masked_code.short_description = 'Штрих-код'
    
    def scanned_by_display(self, obj):
        """Отображает пользователя, который отсканировал."""
        if obj.scanned_by:
            return f"{obj.scanned_by.first_name} (@{obj.scanned_by.username or 'N/A'})"
        return '-'
    scanned_by_display.short_description = 'Пользователь Telegram'
    
    def get_urls(self):
        """Добавляет кастомные URL для генерации QR-кодов."""
        urls = super().get_urls()
        custom_urls = [
            path('generate/', self.admin_site.admin_view(self.generate_qr_codes_view), name='core_qrcode_generate'),
        ]
        return custom_urls + urls
    
    def generate_qr_codes_view(self, request):
        """Представление для генерации QR-кодов."""
        # Проверяем права доступа: только суперадмины могут генерировать QR-коды
        # Call Center не может создавать QR коды
        if not request.user.is_superuser:
            from django.core.exceptions import PermissionDenied
            raise PermissionDenied("У вас нет прав для генерации QR-кодов.")
        
        if request.method == 'POST':
            code_type = request.POST.get('code_type')
            quantity = int(request.POST.get('quantity', 0))
            points = request.POST.get('points')
            
            if code_type and quantity > 0:
                try:
                    # Определяем баллы
                    if points:
                        points = int(points)
                    else:
                        # Используем значения по умолчанию
                        points = settings.ELECTRICIAN_POINTS if code_type == 'electrician' else settings.SELLER_POINTS
                    
                    # Создаем запись о генерации
                    generation = QRCodeGeneration.objects.create(
                        code_type=code_type,
                        quantity=quantity,
                        points=points,
                        created_by=request.user if request.user.is_authenticated else None,
                        status='pending'
                    )
                    
                    # Запускаем Celery задачу
                    from core.tasks import generate_qr_codes_task
                    generate_qr_codes_task.delay(generation.id)
                    
                    messages.success(request, f'Генерация QR-кодов запущена! Вы будете перенаправлены на страницу со списком генераций.')
                    return redirect('admin:core_qrcodegeneration_changelist')
                except Exception as e:
                    messages.error(request, f'Ошибка при запуске генерации: {str(e)}')
            else:
                messages.error(request, 'Заполните все поля корректно!')
        
        # Получаем полный контекст админки (как в dashboard)
        context = {
            **self.admin_site.each_context(request),
            'title': 'Генерация QR-кодов',
            'has_permission': request.user.is_superuser,  # Только superuser может генерировать QR коды
        }
        
        return TemplateResponse(request, 'admin/core/qrcode/generate.html', context)


@admin.register(Gift)
class GiftAdmin(NoDeleteAdminMixin, SimpleHistoryAdmin):
    """Админка для подарков."""
    list_display = ['gift_display', 'user_type_badge', 'points_cost_display', 'order', 'image_preview', 'status_badge', 'created_at']
    list_filter = [
        'is_active', 'user_type',
        ('created_at', DateTimeRangeFilterBuilder(title='Дата создания (диапазон)')),
    ]
    search_fields = ['name_uz_latin', 'name_ru', 'description_uz_latin', 'description_ru']
    readonly_fields = ['created_at', 'updated_at', 'image_preview']
    list_editable = ['order']
    list_per_page = 25
    
    def gift_display(self, obj):
        """Отображает подарок с иконкой."""
        name = obj.name_uz_latin or obj.name_ru or 'Без названия'
        return format_html(
            '<span style="font-size: 20px;">🎁</span> <strong style="font-size: 16px;">{}</strong>',
            name
        )
    gift_display.short_description = 'Подарок'
    gift_display.admin_order_field = 'name_uz_latin'
    
    def user_type_badge(self, obj):
        """Отображает тип пользователя с цветным badge."""
        if obj.user_type == 'electrician':
            return format_html(
                '<span style="background: #fef3c7; color: #92400e; padding: 4px 12px; border-radius: 12px; '
                'font-size: 12px; font-weight: 600;">⚡ Elektrik</span>'
            )
        elif obj.user_type == 'seller':
            return format_html(
                '<span style="background: #dbeafe; color: #1e40af; padding: 4px 12px; border-radius: 12px; '
                'font-size: 12px; font-weight: 600;">🛒 Sotuvchi</span>'
            )
        return format_html(
            '<span style="background: #f3f4f6; color: #6b7280; padding: 4px 12px; border-radius: 12px; '
            'font-size: 12px; font-weight: 600;">🌐 Barcha</span>'
        )
    user_type_badge.short_description = 'Foydalanuvchi turi'
    user_type_badge.admin_order_field = 'user_type'
    
    def points_cost_display(self, obj):
        """Отображает стоимость с цветом."""
        points_formatted = f"{obj.points_cost:,}".replace(",", " ")
        return format_html(
            '<span style="color: #667eea; font-weight: 700; font-size: 16px;">{}</span> баллов',
            points_formatted
        )
    points_cost_display.short_description = 'Стоимость'
    points_cost_display.admin_order_field = 'points_cost'
    
    def status_badge(self, obj):
        """Отображает статус активности."""
        if obj.is_active:
            return format_html(
                '<span style="background: #d4edda; color: #155724; padding: 4px 12px; border-radius: 12px; '
                'font-size: 12px; font-weight: 600;">✅ Активен</span>'
            )
        else:
            return format_html(
                '<span style="background: #f8d7da; color: #721c24; padding: 4px 12px; border-radius: 12px; '
                'font-size: 12px; font-weight: 600;">❌ Неактивен</span>'
            )
    status_badge.short_description = 'Статус'
    status_badge.admin_order_field = 'is_active'
    
    def image_preview(self, obj):
        """Превью изображения подарка."""
        if obj.image:
            return format_html(
                '<img src="{}" style="max-height: 100px; max-width: 100px; border-radius: 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.1);" />',
                obj.image.url
            )
        return '-'
    image_preview.short_description = 'Превью'
    
    fieldsets = (
        ('Основная информация', {
            'fields': ('name_uz_latin', 'name_ru', 'image', 'image_preview')
        }),
        ('Описание', {
            'fields': ('description_uz_latin', 'description_ru')
        }),
        ('Настройки', {
            'fields': ('user_type', 'points_cost', 'order', 'is_active')
        }),
        ('Даты', {
            'fields': ('created_at', 'updated_at')
        }),
    )


@admin.register(GiftRedemption)
class GiftRedemptionAdmin(NoDeleteAdminMixin, SimpleHistoryAdmin):
    """Админка для получения подарков (CRM)."""
    list_display = [
        'redemption_display', 'telegram_id_display', 'phone_number_display', 'region_display',
        'status_badge', 'user_confirmed_badge', 'requested_at'
    ]
    list_filter = [
        'status', 'user_confirmed', 'user__region',
        ('requested_at', DateTimeRangeFilterBuilder(title='Дата запроса (диапазон)')),
    ]
    search_fields = ['user__username', 'user__first_name', 'user__telegram_id', 'user__phone_number', 'gift__name_uz_latin', 'gift__name_ru']
    readonly_fields = ['user', 'gift', 'region_display', 'requested_at', 'confirmed_at']
    list_per_page = 50
    date_hierarchy = 'requested_at'
    
    def redemption_display(self, obj):
        """Отображает информацию о заказе."""
        gift_name = obj.gift.name_uz_latin or obj.gift.name_ru or 'Подарок'
        return format_html(
            '<div style="line-height: 1.6;">'
            '<strong style="font-size: 16px;">🎁 {}</strong><br>'
            '<span style="color: #718096; font-size: 14px;">👤 {}</span>',
            gift_name,
            obj.user.first_name or f"ID: {obj.user.telegram_id}"
        )
    redemption_display.short_description = 'Заказ'
    redemption_display.admin_order_field = 'gift__name_uz_latin'
    
    def telegram_id_display(self, obj):
        """Отображает Telegram ID пользователя."""
        return format_html(
            '<span style="font-family: monospace; color: #3b82f6; font-weight: 600;">{}</span>',
            obj.user.telegram_id
        )
    telegram_id_display.short_description = 'Telegram ID'
    telegram_id_display.admin_order_field = 'user__telegram_id'
    
    def phone_number_display(self, obj):
        """Отображает номер телефона пользователя."""
        if obj.user.phone_number:
            return format_html(
                '<span style="font-family: monospace; color: #10b981; font-weight: 600;">📞 {}</span>',
                obj.user.phone_number
            )
        return format_html('<span style="color: #9ca3af;">-</span>')
    phone_number_display.short_description = 'Телефон'
    phone_number_display.admin_order_field = 'user__phone_number'
    
    def region_display(self, obj):
        """Отображает регион пользователя."""
        region_name = obj.user.get_region_display('ru')
        if region_name:
            return format_html(
                '<span style="background: #e0e7ff; color: #3730a3; padding: 4px 12px; border-radius: 12px; '
                'font-size: 12px; font-weight: 600;">📍 {}</span>',
                region_name
            )
        return format_html('<span style="color: #9ca3af;">-</span>')
    region_display.short_description = 'Регион'
    region_display.admin_order_field = 'user__region'
    
    def status_badge(self, obj):
        """Отображает статус заказа."""
        colors = {
            'pending': ('#fff3cd', '#856404', '⏳'),
            'approved': ('#d4edda', '#155724', '✅'),
            'sent': ('#dbeafe', '#1e40af', '📦'),
            'completed': ('#d1ecf1', '#0c5460', '✔️'),
            'rejected': ('#f8d7da', '#721c24', '❌'),
            'cancelled_by_user': ('#fce4ec', '#c62828', '🚫'),
            'not_received': ('#fff3e0', '#e65100', '⚠️'),
        }
        bg, text, icon = colors.get(obj.status, ('#f3f4f6', '#374151', '📋'))
        label = dict(obj._meta.get_field('status').choices).get(obj.status, obj.status)
        return format_html(
            '<span style="background: {}; color: {}; padding: 4px 12px; border-radius: 12px; '
            'font-size: 12px; font-weight: 600;">{} {}</span>',
            bg, text, icon, label
        )
    status_badge.short_description = 'Статус'
    status_badge.admin_order_field = 'status'
    
    def user_confirmed_badge(self, obj):
        """Отображает подтверждение пользователем."""
        if obj.status == 'not_received':
            return format_html(
                '<span style="background: #fee2e2; color: #dc2626; padding: 4px 12px; border-radius: 12px; '
                'font-size: 12px; font-weight: 600;">❌ Подарок не выдан</span>'
            )
        elif obj.user_confirmed is True:
            return format_html(
                '<span style="background: #d4edda; color: #155724; padding: 4px 12px; border-radius: 12px; '
                'font-size: 12px; font-weight: 600;">✅ Подтверждено</span>'
            )
        elif obj.user_confirmed is False and obj.status != 'not_received':
            return format_html(
                '<span style="background: #fff3cd; color: #856404; padding: 4px 12px; border-radius: 12px; '
                'font-size: 12px; font-weight: 600;">⚠️ Не подтверждено</span>'
            )
        return format_html(
            '<span style="background: #f3f4f6; color: #6b7280; padding: 4px 12px; border-radius: 12px; '
            'font-size: 12px; font-weight: 600;">-</span>'
        )
    user_confirmed_badge.short_description = 'Подтверждение'
    user_confirmed_badge.admin_order_field = 'user_confirmed'
    
    fieldsets = (
        ('Информация о запросе', {
            'fields': ('user', 'gift', 'region_display', 'requested_at')
        }),
        ('Обработка', {
            'fields': ('status', 'admin_notes')
        }),
        ('Подтверждение пользователем', {
            'fields': ('user_confirmed', 'user_comment', 'confirmed_at')
        }),
    )
    
    def formfield_for_dbfield(self, db_field, request, **kwargs):
        """Ограничивает выбор статусов в зависимости от роли пользователя."""
        if db_field.name == 'status':
            # Проверяем, имеет ли пользователь пермишн агента
            is_agent = request.user.has_perm('core.change_status_agent')
            # Проверяем, имеет ли пользователь пермишн call center
            is_call_center = request.user.has_perm('core.change_user_type_call_center')
            
            if not request.user.is_superuser:
                choices = list(GiftRedemption.STATUS_CHOICES)
                
                # Агенты видят только 'sent' и 'completed'
                if is_agent:
                    filtered_choices = [choice for choice in choices if choice[0] in ['sent', 'completed']]
                    kwargs['choices'] = filtered_choices
                # Call Center видит только 'pending', 'approved', 'sent'
                elif is_call_center:
                    filtered_choices = [choice for choice in choices if choice[0] in 
                    ['pending', 'approved', 'sent', 'completed', 'rejected', 'cancelled_by_user']
                    ]
                kwargs['choices'] = filtered_choices
            
        return super().formfield_for_dbfield(db_field, request, **kwargs)
    
    def get_readonly_fields(self, request, obj=None):
        """Управляет readonly полями в зависимости от роли пользователя."""
        readonly = list(super().get_readonly_fields(request, obj))
        
        if not request.user.is_superuser:
            # Агенты могут изменять только status
            if request.user.has_perm('core.change_status_agent'):
                # Получаем все поля модели
                model_fields = [
                    f.name for f in GiftRedemption._meta.get_fields() 
                    if isinstance(f, models.Field) and hasattr(f, 'name')
                ]
                # Делаем все поля readonly кроме status
                for field in model_fields:
                    if field != 'status' and field not in readonly:
                        readonly.append(field)
            # Call Center не может подтверждать получение подарка
            elif request.user.has_perm('core.change_user_type_call_center'):
                if 'user_confirmed' not in readonly:
                    readonly.append('user_confirmed')
        
        return readonly
    
    def save_model(self, request, obj, form, change):
        """Автоматически устанавливает processed_at при изменении статуса и отправляет уведомления."""
        # Проверяем права доступа
        is_agent = request.user.has_perm('core.change_status_agent')
        is_call_center = request.user.has_perm('core.change_user_type_call_center')
        
        # Call Center не может устанавливать статус 'completed' (но агенты могут)
        if is_call_center and not is_agent and not request.user.is_superuser:
            if obj.status == 'completed':
                from django.core.exceptions import PermissionDenied
                raise PermissionDenied(
                    "Сотрудники call center не могут устанавливать статус 'completed' "
                    "(клиент подтвердил получение заказа). Этот статус может быть установлен только клиентом или агентами."
                )
        
        old_status = None
        
        if change:
            # Получаем старые значения статуса
            old_obj = GiftRedemption.objects.get(pk=obj.pk)
            old_status = old_obj.status
        
        # Сохраняем объект
        super().save_model(request, obj, form, change)
        
        # Инвалидируем кеш баллов при изменении статуса на отмену/отклонение
        if change and 'status' in form.changed_data:
            if obj.status in ['rejected', 'cancelled_by_user', 'not_received']:
                obj.user.invalidate_points_cache()
                obj.user.calculate_points(force=True)
        
        # Отправляем уведомления после сохранения
        if change:
            import asyncio
            from aiogram import Bot
            from bot.translations import get_text
            
            async def send_notification():
                try:
                    bot = Bot(token=settings.TELEGRAM_BOT_TOKEN)
                    user = obj.user
                    gift_name = obj.gift.get_name(user.language if user else 'uz_latin')
                    
                    # Уведомление об изменении статуса
                    if 'status' in form.changed_data and old_status != obj.status:
                        if obj.status == 'approved':
                            message = get_text(user, 'GIFT_STATUS_APPROVED', gift_name=gift_name)
                        elif obj.status == 'sent':
                            message = get_text(user, 'GIFT_STATUS_SENT', gift_name=gift_name)
                        elif obj.status == 'completed':
                            message = get_text(user, 'GIFT_STATUS_COMPLETED', gift_name=gift_name)
                        elif obj.status == 'rejected':
                            admin_notes = obj.admin_notes or ""
                            # Формируем текст причины в зависимости от языка пользователя
                            if user.language == 'ru':
                                if admin_notes and admin_notes.strip():
                                    admin_notes_text = f"Причина: {admin_notes}"
                                else:
                                    admin_notes_text = "Причина не указана"
                            else:  # uz_latin
                                if admin_notes and admin_notes.strip():
                                    admin_notes_text = f"Sabab: {admin_notes}"
                                else:
                                    admin_notes_text = "Sabab ko'rsatilmagan"
                            message = get_text(user, 'GIFT_STATUS_REJECTED', gift_name=gift_name, admin_notes=admin_notes_text)
                        else:
                            message = None
                        
                        if message:
                            from core.messaging import send_message_to_user
                            await send_message_to_user(bot, user, message)
                    
                    await bot.session.close()
                except Exception as e:
                    import logging
                    logger = logging.getLogger(__name__)
                    logger.error(f"Ошибка при отправке уведомления о статусе подарка: {e}")
            
            # Запускаем асинхронную функцию в отдельном потоке
            # Это необходимо, так как Django admin работает в синхронном контексте
            import threading
            
            def run_async_in_thread():
                """Запускает асинхронную функцию в новом event loop в отдельном потоке."""
                new_loop = asyncio.new_event_loop()
                asyncio.set_event_loop(new_loop)
                try:
                    new_loop.run_until_complete(send_notification())
                finally:
                    new_loop.close()
            
            thread = threading.Thread(target=run_async_in_thread, daemon=True)
            thread.start()


@admin.register(RegionMessageLog)
class RegionMessageLogAdmin(NoDeleteAdminMixin, admin.ModelAdmin):
    """Логи рассылок по областям (результаты Celery-задач)."""
    list_display = [
        'region_code', 'total', 'sent_count', 'failed_count', 'status',
        'initiated_by', 'created_at', 'completed_at',
    ]
    list_filter = ['status', 'region_code', ('created_at', DateTimeRangeFilterBuilder(title='Дата'))]
    readonly_fields = [
        'region_code', 'user_type_filter', 'language_filter',
        'total', 'sent_count', 'failed_count', 'status',
        'initiated_by', 'created_at', 'completed_at', 'error_message',
    ]
    search_fields = ['region_code', 'error_message']
    ordering = ['-created_at']
    date_hierarchy = 'created_at'

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False


@admin.register(BroadcastMessage)
class BroadcastMessageAdmin(NoDeleteAdminMixin, SimpleHistoryAdmin):
    """Админка для массовых рассылок."""
    list_display = [
        'title', 'status', 'user_type_filter', 'total_users',
        'sent_count', 'failed_count', 'created_at', 'completed_at', 'send_button'
    ]
    list_filter = [
        'status', 'user_type_filter', 'region_filter',
        ('created_at', DateTimeRangeFilterBuilder(title='Дата создания (диапазон)')),
    ]
    search_fields = ['title', 'message_text']
    readonly_fields = [
        'status', 'total_users', 'sent_count', 'failed_count',
        'created_at', 'started_at', 'completed_at'
    ]
    
    fieldsets = (
        ('Основная информация', {
            'fields': ('title', 'message_text', 'image', 'user_type_filter'),
            'description': 'Текст поддерживает HTML: <b>жирный</b>, <i>курсив</i>, <a href="url">ссылка</a>. Эмодзи и стикеры можно вставлять в текст. Фото — опционально.'
        }),
        ('Фильтрация по региону', {
            'fields': ('region_filter',),
            'description': 'Выберите область для фильтрации пользователей. Если не выбрано, сообщение будет отправлено всем пользователям.'
        }),
        ('Статистика', {
            'fields': (
                'status', 'total_users', 'sent_count', 'failed_count',
                'created_at', 'started_at', 'completed_at'
            )
        }),
    )
    
    actions = ['send_broadcast_action']
    
    def send_button(self, obj):
        """Кнопка отправки рассылки в списке."""
        if obj.status == 'pending':
            from django.urls import reverse
            url = reverse('admin:core_broadcastmessage_send_single', args=[obj.pk])
            return format_html(
                '<a href="{}" style="background: #28a745; color: white; padding: 6px 12px; '
                'border-radius: 4px; text-decoration: none; white-space: nowrap; font-size: 12px;" '
                'onclick="return confirm(\'Отправить рассылку?\');">📤 Отправить</a>',
                url
            )
        elif obj.status == 'sending':
            return format_html(
                '<span style="color: #1e40af; font-size: 12px;">🔄 Отправляется...</span>'
            )
        elif obj.status == 'completed':
            return format_html(
                '<span style="color: #155724; font-size: 12px;">✅ Отправлено</span>'
            )
        return '-'
    send_button.short_description = 'Действие'
    
    def get_urls(self):
        """Добавляет кастомные URL."""
        urls = super().get_urls()
        custom_urls = [
            path('<int:broadcast_id>/send/', self.admin_site.admin_view(self.send_single_broadcast_view), name='core_broadcastmessage_send_single'),
        ]
        return custom_urls + urls
    
    def send_single_broadcast_view(self, request, broadcast_id):
        """Отправка конкретной рассылки."""
        import subprocess
        
        try:
            broadcast = BroadcastMessage.objects.get(pk=broadcast_id)
        except BroadcastMessage.DoesNotExist:
            self.message_user(request, 'Рассылка не найдена', messages.ERROR)
            return redirect('admin:core_broadcastmessage_changelist')
        
        if broadcast.status != 'pending':
            self.message_user(request, f'Рассылка "{broadcast.title}" уже была отправлена', messages.WARNING)
            return redirect('admin:core_broadcastmessage_changelist')
        
        # Предварительно оцениваем количество пользователей
        users_query = TelegramUser.objects.filter(is_active=True)
        if broadcast.user_type_filter:
            users_query = users_query.filter(user_type=broadcast.user_type_filter)
        estimated_users = users_query.count()
        
        LARGE_BROADCAST_THRESHOLD = 20000
        
        if estimated_users >= LARGE_BROADCAST_THRESHOLD:
            try:
                from core.tasks import send_broadcast_chained
                send_broadcast_chained.delay(broadcast.id)
                self.message_user(request, f'Рассылка "{broadcast.title}" запущена через Celery ({estimated_users} пользователей)', messages.SUCCESS)
            except Exception as e:
                self.message_user(request, f'Ошибка: {e}', messages.ERROR)
        else:
            try:
                subprocess.Popen(['python', 'manage.py', 'send_broadcast', str(broadcast.id)])
                self.message_user(request, f'Рассылка "{broadcast.title}" запущена', messages.SUCCESS)
            except Exception as e:
                self.message_user(request, f'Ошибка: {e}', messages.ERROR)
        
        return redirect('admin:core_broadcastmessage_changelist')
    
    def formfield_for_dbfield(self, db_field, request, **kwargs):
        """Добавляет опцию 'Всем' в фильтр по типу пользователя."""
        if db_field.name == 'user_type_filter':
            # Получаем текущие choices из модели
            from core.models import TelegramUser
            choices = list(TelegramUser.USER_TYPE_CHOICES)
            # Добавляем опцию "Всем" в начало списка
            choices.insert(0, ('', 'Всем'))
            kwargs['choices'] = choices
            kwargs['required'] = False
        return super().formfield_for_dbfield(db_field, request, **kwargs)
    
    def send_broadcast_action(self, request, queryset):
        """Действие для отправки рассылки."""
        import subprocess
        from django.contrib import messages
        
        for broadcast in queryset:
            if broadcast.status != 'pending':
                self.message_user(
                    request,
                    f'Рассылка "{broadcast.title}" уже была отправлена',
                    level=messages.WARNING
                )
                continue
            
            # Определяем, использовать ли Celery для больших рассылок
            LARGE_BROADCAST_THRESHOLD = 20000  # Порог для использования Celery
            
            # Предварительно оцениваем количество пользователей
            from core.models import TelegramUser
            users_query = TelegramUser.objects.filter(is_active=True)
            if broadcast.user_type_filter:
                users_query = users_query.filter(user_type=broadcast.user_type_filter)
            
            estimated_users = users_query.count()
            
            # Если пользователей много, используем Celery
            if estimated_users >= LARGE_BROADCAST_THRESHOLD:
                try:
                    from core.tasks import send_broadcast_chained
                    send_broadcast_chained.delay(broadcast.id)
                    self.message_user(
                        request,
                        f'Рассылка "{broadcast.title}" запущена через Celery ({estimated_users} пользователей)',
                        level=messages.SUCCESS
                    )
                except Exception as e:
                    self.message_user(
                        request,
                        f'Ошибка при запуске рассылки через Celery: {e}',
                        level=messages.ERROR
                    )
            else:
                # Для небольших рассылок используем обычную команду
                try:
                    subprocess.Popen([
                        'python', 'manage.py', 'send_broadcast', str(broadcast.id)
                    ])
                    self.message_user(
                        request,
                        f'Рассылка "{broadcast.title}" запущена',
                        level=messages.SUCCESS
                    )
                except Exception as e:
                    self.message_user(
                        request,
                        f'Ошибка при запуске рассылки: {e}',
                        level=messages.ERROR
                    )
    send_broadcast_action.short_description = 'Отправить выбранные рассылки'


@admin.register(Promotion)
class PromotionAdmin(NoDeleteAdminMixin, SimpleHistoryAdmin):
    """Админка для акций/баннеров."""
    list_display = [
        'image_preview', 'title', 'date_display', 'order', 'is_active', 'status_badge', 'created_at'
    ]
    list_filter = [
        'is_active',
        ('created_at', DateTimeRangeFilterBuilder(title='Дата создания (диапазон)')),
        ('date', DateRangeFilterBuilder(title='Дата акции (диапазон)')),
    ]
    search_fields = ['title']
    list_editable = ['order', 'is_active']
    ordering = ['order', '-created_at']
    date_hierarchy = 'created_at'
    
    fieldsets = (
        ('Основная информация', {
            'fields': ('title', 'image', 'date', 'order', 'is_active')
        }),
        ('Системная информация', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )
    
    readonly_fields = ['created_at', 'updated_at']
    
    def image_preview(self, obj):
        """Превью изображения акции."""
        if obj.image:
            return format_html(
                '<img src="{}" style="max-width: 100px; max-height: 100px; object-fit: cover; border-radius: 8px;" />',
                obj.image.url
            )
        return '-'
    image_preview.short_description = 'Rasm'
    
    def date_display(self, obj):
        """Отображает дату в формате DD.MM.YYYY."""
        if obj.date:
            return obj.date.strftime('%d.%m.%Y')
        return '-'
    date_display.short_description = 'Sana'
    date_display.admin_order_field = 'date'
    
    def status_badge(self, obj):
        """Отображает статус активности."""
        if obj.is_active:
            return format_html(
                '<span style="background: #10B981; color: white; padding: 4px 8px; border-radius: 4px; font-size: 11px;">Faol</span>'
            )
        return format_html(
            '<span style="background: #EF4444; color: white; padding: 4px 8px; border-radius: 4px; font-size: 11px;">Nofaol</span>'
        )
    status_badge.short_description = 'Holat'


@admin.register(QRCodeGeneration)
class QRCodeGenerationAdmin(NoDeleteAdminMixin, SimpleHistoryAdmin):
    """Админка для истории генерации QR-кодов."""
    list_display = [
        'generation_display', 'code_type_badge', 'quantity_display',
        'points_display', 'status_badge', 'created_by_display',
        'created_at', 'completed_at_display', 'download_button'
    ]
    list_filter = [
        'status', 'code_type',
        ('created_at', DateTimeRangeFilterBuilder(title='Дата создания (диапазон)')),
    ]
    search_fields = ['id']
    readonly_fields = [
        'code_type', 'quantity', 'points', 'status', 'zip_file',
        'qr_codes', 'error_message', 'created_by', 'created_at', 'completed_at'
    ]
    ordering = ['-created_at']
    list_per_page = 50
    date_hierarchy = 'created_at'
    
    def generation_display(self, obj):
        """Отображает информацию о генерации."""
        return format_html(
            '<div style="line-height: 1.6;">'
            '<strong style="font-size: 16px;">#{}</strong><br>'
            '<span style="color: #718096; font-size: 12px;">{} шт.</span>',
            obj.id,
            obj.quantity
        )
    generation_display.short_description = 'Генерация'
    generation_display.admin_order_field = 'id'
    
    def code_type_badge(self, obj):
        """Отображает тип кода."""
        if obj.code_type == 'electrician':
            return format_html(
                '<span style="background: #fef3c7; color: #92400e; padding: 4px 12px; border-radius: 12px; '
                'font-size: 12px; font-weight: 600;">⚡ E-</span>'
            )
        elif obj.code_type == 'seller':
            return format_html(
                '<span style="background: #dbeafe; color: #1e40af; padding: 4px 12px; border-radius: 12px; '
                'font-size: 12px; font-weight: 600;">🛒 D-</span>'
            )
        return '-'
    code_type_badge.short_description = 'Тип'
    code_type_badge.admin_order_field = 'code_type'
    
    def quantity_display(self, obj):
        """Отображает количество."""
        return format_html(
            '<span style="font-weight: 600;">{}</span>',
            obj.quantity
        )
    quantity_display.short_description = 'Количество'
    quantity_display.admin_order_field = 'quantity'
    
    def points_display(self, obj):
        """Отображает баллы."""
        return format_html(
            '<span style="color: #667eea; font-weight: 700;">{} баллов</span>',
            obj.points
        )
    points_display.short_description = 'Баллы'
    points_display.admin_order_field = 'points'
    
    def status_badge(self, obj):
        """Отображает статус генерации."""
        colors = {
            'pending': ('#fff3cd', '#856404', '⏳'),
            'processing': ('#dbeafe', '#1e40af', '🔄'),
            'completed': ('#d4edda', '#155724', '✅'),
            'failed': ('#f8d7da', '#721c24', '❌'),
        }
        bg, text, icon = colors.get(obj.status, ('#f3f4f6', '#374151', '📋'))
        label = dict(obj._meta.get_field('status').choices).get(obj.status, obj.status)
        return format_html(
            '<span style="background: {}; color: {}; padding: 4px 12px; border-radius: 12px; '
            'font-size: 12px; font-weight: 600;">{} {}</span>',
            bg, text, icon, label
        )
    status_badge.short_description = 'Статус'
    status_badge.admin_order_field = 'status'
    
    def created_by_display(self, obj):
        """Отображает создателя."""
        if obj.created_by:
            return obj.created_by.username or str(obj.created_by)
        return '-'
    created_by_display.short_description = 'Создал'
    
    def completed_at_display(self, obj):
        """Отображает время завершения."""
        if obj.completed_at:
            return obj.completed_at.strftime('%d.%m.%Y %H:%M')
        return '-'
    completed_at_display.short_description = 'Завершено'
    completed_at_display.admin_order_field = 'completed_at'
    
    def download_button(self, obj):
        """Кнопки для скачивания ZIP файла и Excel."""
        if obj.status == 'completed':
            buttons = []
            if obj.zip_file:
                buttons.append(
                    '<a href="{}" style="background: #417690; color: white; padding: 6px 12px; '
                    'border-radius: 4px; text-decoration: none; display: inline-block; margin-right: 5px;">📥 .zip</a>'.format(
                        obj.zip_file.url
                    )
                )
            buttons.append(
                '<a href="{}" style="background: #28a745; color: white; padding: 6px 12px; '
                'border-radius: 4px; text-decoration: none; display: inline-block;">📊 .xlsx</a>'.format(
                    f'/admin/core/qrcodegeneration/{obj.id}/export_excel/'
                )
            )
            return format_html(''.join(buttons))
        elif obj.status == 'failed':
            return format_html(
                '<span style="color: #dc3545; font-size: 11px;">{}</span>',
                obj.error_message[:50] + '...' if obj.error_message and len(obj.error_message) > 50 else obj.error_message or 'Ошибка'
            )
        return '-'
    download_button.short_description = 'Действие'
    
    fieldsets = (
        ('Основная информация', {
            'fields': ('code_type', 'quantity', 'points', 'status')
        }),
        ('Результаты', {
            'fields': ('zip_file', 'qr_codes', 'error_message')
        }),
        ('Системная информация', {
            'fields': ('created_by', 'created_at', 'completed_at')
        }),
    )
    
    def has_module_permission(self, request):
        """Только superuser может видеть этот модуль в меню."""
        return request.user.is_superuser
    
    def has_add_permission(self, request):
        """Отключаем добавление через админку."""
        return False

    def get_urls(self):
        """Добавляет кастомные URL для экспорта Excel."""
        urls = super().get_urls()
        custom_urls = [
            path('<path:object_id>/export_excel/', self.admin_site.admin_view(self.export_excel_view), name='core_qrcodegeneration_export_excel'),
        ]
        return custom_urls + urls
    
    def export_excel_view(self, request, object_id):
        """Экспорт QR-кодов в Excel формат."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from django.utils import timezone
        
        try:
            generation = QRCodeGeneration.objects.get(id=object_id)
        except QRCodeGeneration.DoesNotExist:
            from django.http import Http404
            raise Http404("Генерация не найдена")
        
        # Получаем все QR-коды для этой генерации
        qr_codes = generation.qr_codes.all().order_by('generated_at')
        
        # Создаем Excel файл
        wb = Workbook()
        ws = wb.active
        ws.title = "QR Codes"
        
        # Заголовки
        headers = ['Дата создания QR кода', 'Серийный номер', 'Сканирован ли', 'Промо код']
        ws.append(headers)
        
        # Стили для заголовков
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center')
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Данные
        for qr_code in qr_codes:
            generated_at = qr_code.generated_at.strftime('%d.%m.%Y %H:%M:%S') if qr_code.generated_at else ''
            serial_number = qr_code.serial_number
            is_scanned = 'Да' if qr_code.is_scanned else 'Нет'
            promo_code = qr_code.code
            ws.append([generated_at, serial_number, is_scanned, promo_code])
        
        # Автоподбор ширины колонок
        from openpyxl.utils import get_column_letter
        column_widths = [25, 20, 15, 20]
        for idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = width
        
        # Создаем HttpResponse с Excel файлом
        filename = f"qrcodes_{generation.id}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response


@admin.register(PrivacyPolicy)
class PrivacyPolicyAdmin(NoDeleteAdminMixin, SimpleHistoryAdmin):
    """Админка для политики конфиденциальности."""
    list_display = ['is_active', 'updated_at', 'created_at', 'has_pdf_files']
    list_display_links = ['is_active', 'updated_at', 'created_at', 'has_pdf_files']
    list_filter = [
        'is_active',
    ]
    fieldsets = (
        ('Узбекский язык (Латиница)', {
            'fields': ('pdf_uz_latin',),
        }),
        ('Русский язык', {
            'fields': ('pdf_ru',),
        }),
        ('Настройки', {
            'fields': ('is_active',),
        }),
        ('Системная информация', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )
    readonly_fields = ['created_at', 'updated_at']
    
    def has_pdf_files(self, obj):
        """Показывает, загружены ли PDF файлы."""
        if not obj:
            return '-'
        pdfs = []
        if obj.pdf_uz_latin:
            pdfs.append('UZ (Lat)')
        if obj.pdf_ru:
            pdfs.append('RU')
        return ', '.join(pdfs) if pdfs else 'Нет PDF'
    has_pdf_files.short_description = 'Загруженные PDF'
    
    def has_add_permission(self, request):
        """Разрешаем создание только для superuser (Call Center не может создавать QR коды)."""
        # Только superuser может создавать QR коды
        # Call Center не может создавать QR коды, даже если у них есть permission generate_qrcodes
        return request.user.is_superuser


@admin.register(AdminContactSettings)
class AdminContactSettingsAdmin(NoDeleteAdminMixin, SimpleHistoryAdmin):
    """Админка для настроек контакта администратора."""
    list_display = ['contact_type_display', 'contact_value_display', 'is_active', 'updated_at']
    list_filter = [
        'contact_type', 'is_active',
        ('updated_at', DateTimeRangeFilterBuilder(title='Дата обновления (диапазон)')),
    ]
    search_fields = ['contact_value']
    fields = ['contact_type', 'contact_value', 'is_active']
    readonly_fields = ['created_at', 'updated_at']
    
    def contact_type_display(self, obj):
        """Отображает тип контакта с иконкой."""
        icons = {
            'telegram': '💬',
            'phone': '📞',
            'link': '🔗',
        }
        icon = icons.get(obj.contact_type, '📋')
        return format_html('{} {}', icon, obj.get_contact_type_display())
    contact_type_display.short_description = 'Kontakt turi'
    
    def contact_value_display(self, obj):
        """Отображает значение контакта с предпросмотром URL."""
        url = obj.get_contact_url()
        if url:
            return format_html(
                '<strong>{}</strong><br><a href="{}" target="_blank" style="color: #2064AE; font-size: 12px;">{}</a>',
                obj.contact_value, url, url
            )
        return obj.contact_value
    contact_value_display.short_description = 'Kontakt qiymati'
    
    def has_add_permission(self, request):
        """Разрешаем создание только для superuser."""
        return request.user.is_superuser
    
    def save_model(self, request, obj, form, change):
        """При сохранении деактивируем другие активные настройки, если эта активна."""
        if obj.is_active:
            # Деактивируем все другие активные настройки
            AdminContactSettings.objects.filter(is_active=True).exclude(pk=obj.pk if obj.pk else None).update(is_active=False)
        super().save_model(request, obj, form, change)


@admin.register(VideoInstruction)
class VideoInstructionAdmin(NoDeleteAdminMixin, SimpleHistoryAdmin):
    """Админка для видео инструкций. 4 видео: электрики (UZ/RU) и предприниматели (UZ/RU)."""
    list_display = ['video_electrician_preview', 'video_seller_preview', 'file_id_status', 'is_active', 'updated_at']
    list_filter = [
        'is_active',
        ('updated_at', DateTimeRangeFilterBuilder(title='Дата обновления (диапазон)')),
    ]
    fieldsets = (
        ('Video — Elektriklar', {
            'fields': ('video_electrician_uz', 'thumb_electrician_uz', 'video_electrician_ru', 'thumb_electrician_ru')
        }),
        ('Video — Tadbirkorlar', {
            'fields': ('video_seller_uz', 'thumb_seller_uz', 'video_seller_ru', 'thumb_seller_ru')
        }),
        ('Telegram file_id (avtomatik)', {
            'fields': ('file_id_electrician_uz', 'file_id_electrician_ru', 'file_id_seller_uz', 'file_id_seller_ru'),
            'description': 'File_id avtomatik to\'ldiriladi. Thumbnail — JPEG max 320x320, 200KB (oldindan ko\'rinish uchun).'
        }),
        ('Sozlamalar', {
            'fields': ('is_active',)
        }),
        ('Sana', {
            'fields': ('created_at', 'updated_at')
        }),
    )
    readonly_fields = ['created_at', 'updated_at', 'file_id_electrician_uz', 'file_id_electrician_ru', 'file_id_seller_uz', 'file_id_seller_ru']
    
    def video_electrician_preview(self, obj):
        uz = '✅' if obj.video_electrician_uz else '❌'
        ru = '✅' if obj.video_electrician_ru else '❌'
        return format_html('<span>⚡ Elektrik: UZ {} | RU {}</span>', uz, ru)
    video_electrician_preview.short_description = 'Video (Elektrik)'
    
    def video_seller_preview(self, obj):
        uz = '✅' if obj.video_seller_uz else '❌'
        ru = '✅' if obj.video_seller_ru else '❌'
        return format_html('<span>🛒 Tadbirkor: UZ {} | RU {}</span>', uz, ru)
    video_seller_preview.short_description = 'Video (Tadbirkor)'
    
    def file_id_status(self, obj):
        return format_html(
            '⚡ UZ{} RU{} | 🛒 UZ{} RU{}',
            '✅' if obj.file_id_electrician_uz else '❌',
            '✅' if obj.file_id_electrician_ru else '❌',
            '✅' if obj.file_id_seller_uz else '❌',
            '✅' if obj.file_id_seller_ru else '❌',
        )
    file_id_status.short_description = 'File ID'
    
    def has_add_permission(self, request):
        """Разрешаем создание только для superuser."""
        return request.user.is_superuser
    
    def save_model(self, request, obj, form, change):
        """При сохранении деактивируем другие активные инструкции, если эта активна."""
        if obj.is_active:
            # Деактивируем все другие активные инструкции
            VideoInstruction.objects.filter(is_active=True).exclude(pk=obj.pk if obj.pk else None).update(is_active=False)
        super().save_model(request, obj, form, change)


@admin.register(SmartUPId)
class SmartUPIdAdmin(NoDeleteAdminMixin, admin.ModelAdmin):
    """Админка для SmartUP ID."""
    list_display = ['id_value', 'created_at']
    search_fields = ['id_value']
    ordering = ['id_value']
    readonly_fields = ['created_at']
    list_per_page = 100
    
    fieldsets = (
        ('Основная информация', {
            'fields': ('id_value', 'created_at')
        }),
    )


# Кастомная админка для дашборда
admin.site.site_header = 'Mona Admin Panel'
admin.site.site_title = 'Mona Admin'
admin.site.index_title = 'Панель управления'


"""
EAAbotoqNT04BQyNhjTxdjGaRZCRuGyeObZBs1w3DKYbZAj6TeCO698iFKZCn2A2FpnHmoz1FLYZBxQkhJ6ruZC7R0hNOXSmnuZBBiczhTitX3EEJe3nNgzZBsyShLcAa6OOHCjaZAM6E61uwzmpVWAMcwiFtKCqgdevFPh6IxG6F9BFXQKzY1ZBI3kGxjwexcHfCk72glGEyZApp9s0zEQZCu4fBxhE6dBrWZCtZCZAfgLIgU63vOT09Iy9lBQRYKpbmc0MthPZBza6FR3FCc0WMeON2wZCZCtUr8K
"""